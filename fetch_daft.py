
# @author: Miguel Montero Montiano
# @date: 23/09/2017
# @title: Data extraction from Daft alerts and Excel file update

from gmail import Gmail
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

# Functions section

# Function: Convert str to date

def create_date(date_str):
	year = int(date_str.split("-")[0])
	month = int(date_str.split("-")[1])
	day = int(date_str.split("-")[2])

	date_formated = datetime.date(year,month,day)
	return date_formated

# End functions 

g = Gmail()

# Find a way to 'hide' your credentials. That must be insecure :) 
g.login("your_email", "your_password")

alerts = g.inbox().mail(prefetch=True, unread=True, sender='noreply@daft.ie')

extractions = []

for alert in alerts:
    if alert.body.find("Photos") != -1:
	    address = alert.body.split("We found a new property for you:")[1].split("https")[0].strip()
	    area = address.split(",")[-2].strip()
	    acc = alert.body.split("Photos")[1].split("<strong")[0].strip()
	    price = alert.body.split("&euro;")[1].split("</strong")[0]
	    price = int(price. replace(",", ""))
	    per_mth_week = alert.body.split("</strong>")[1].split("To opt out")[0].strip()
	    beds = acc.split("|")[1].strip()
	    baths = acc.split("|")[2].strip()
	    type_acc = acc.split("|")[0]
	    ext = {'date': create_date(str(alert.sent_at).split()[0]), 'time': str(alert.sent_at).split()[1], 
	    'address':address, 'area': area, 'type': type_acc, 'beds': beds, 'baths': baths, 
	    'price': price, 'per_mth_week': per_mth_week}
	    extractions.append(ext)


print ("\n" + "Extraction has been completed!")
print("\n" + "Alerts in INBOX: %d" %len(alerts))
print("\n" + "Alerts extracted: %d" %len(extractions))

# Clean up INBOX 

for alert in alerts:
	alert.delete()

# Transfer data to Excel

wb = Workbook()
wb = load_workbook('daft_alerts_data.xlsx')
ws = wb['data']

for data in extractions:
	row = [data['date'], data['time'], data['address'], data['area'], data['type'], data['beds'], 
	data['baths'], data['price'], data['per_mth_week']]
	ws.append(row)

wb.save("daft_alerts_data.xlsx") 

# Close session

g.logout()
